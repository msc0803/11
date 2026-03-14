"""
视频素材批量查询 & 加入工作台 - PyQt6 GUI
"""
import sys
import os
import time
import math
import pandas as pd
from datetime import datetime

def _read_numbers_file(path):
    """读取 Apple Numbers 文件，返回 {sheet_name: DataFrame}"""
    import numbers_parser
    doc = numbers_parser.Document(path)
    result = {}
    for sheet in doc.sheets:
        rows = sheet.tables[0].rows(values_only=True)
        rows = list(rows)
        if not rows:
            result[sheet.name] = pd.DataFrame()
            continue
        df = pd.DataFrame(rows[1:], columns=rows[0])
        result[sheet.name] = df
    return result

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QFileDialog, QComboBox, QTableWidget,
    QTableWidgetItem, QTextEdit, QProgressBar, QGroupBox, QSpinBox,
    QSplitter, QHeaderView, QMessageBox, QStatusBar, QLineEdit, QCheckBox, QMenu
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QSortFilterProxyModel
from PyQt6.QtGui import QColor, QFont

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import api_core

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
os.makedirs(RESULTS_DIR, exist_ok=True)


# ─────────────────────────── 后台线程 ────────────────────────────

class TokenWorker(QThread):
    success = pyqtSignal(str)
    failed  = pyqtSignal(str)
    log     = pyqtSignal(str)

    def __init__(self, ask_login=False):
        super().__init__()
        self.ask_login = ask_login

    def run(self):
        try:
            if not self.ask_login:
                token = api_core.load_cached_token()
                if token and api_core.is_token_valid(token):
                    self.success.emit(token)
                    return
            self.log.emit("正在启动浏览器...")
            api_core._ensure_chrome_debug()
            self.log.emit("浏览器已就绪")
            token = api_core.get_token_from_browser()
            api_core.save_token(token)
            self.success.emit(token)
        except Exception as e:
            self.failed.emit(str(e))


class SearchWorker(QThread):
    progress     = pyqtSignal(int, int)
    row_done     = pyqtSignal(int, str, str, str, str, str)
    log          = pyqtSignal(str)
    finished_signal = pyqtSignal(int, int, int, int)
    token_refreshed = pyqtSignal(str)

    def __init__(self, df, name_col, start_row, token, max_count=60):
        super().__init__()
        self.df        = df
        self.name_col  = name_col
        self.start_row = start_row
        self.token     = token
        self.max_count = max_count
        self._stop     = False

    def stop(self): self._stop = True

    def run(self):
        found = not_found = error = skip = 0
        total = len(self.df)
        done  = 0

        for idx, row in self.df.iterrows():
            if self._stop:
                self.log.emit("用户停止查询")
                break

            # 跳过逻辑
            if idx < self.start_row:
                skip += 1; done += 1
                self.progress.emit(done, total); continue

            name = row[self.name_col]
            if pd.isna(name) or str(name).strip() == "":
                skip += 1; done += 1
                self.progress.emit(done, total); continue
            name = str(name).strip()

            if str(row.get("查询状态", "")).strip() == "找到":
                skip += 1; done += 1
                self.progress.emit(done, total); continue

            # 查询 + 限流重试
            self.log.emit(f"查询 [{done+1}/{total}]: {name}")
            ok, vid, extra = api_core.search_video(self.token, name, self.max_count)
            for _ in range(3):
                if extra.get("error", "").startswith("当前访问人数"):
                    time.sleep(3)
                    ok, vid, extra = api_core.search_video(self.token, name, self.max_count)
                else:
                    break

            # token 过期
            if extra.get("token_expired"):
                self.log.emit("Token 失效，自动刷新...")
                try:
                    self.token = api_core.get_token_from_browser()
                    api_core.save_token(self.token)
                    self.token_refreshed.emit(self.token)
                    ok, vid, extra = api_core.search_video(self.token, name, self.max_count)
                except Exception as e:
                    self.log.emit(f"Token 刷新失败: {e}")

            if ok:
                cnt    = extra.get("match_count", 1)
                raws   = extra.get("all_raws", [])
                names  = "；".join(v.get("name", "") for v in raws)
                remark = f"匹配{cnt}条 | state={extra.get('videoState','')} cost={extra.get('sumStatCost',0)} roi={extra.get('sumRoi',0)}"
                self.log.emit(f"  ✅ 找到 {cnt} 个视频: {vid}")
                self.row_done.emit(idx, name, "找到", vid, names, remark)
                found += 1
            elif "error" in extra:
                self.log.emit(f"  ❌ 错误: {extra['error']}")
                self.row_done.emit(idx, name, "错误", "", "", extra["error"])
                error += 1
            else:
                self.log.emit(f"  — 未找到")
                self.row_done.emit(idx, name, "未找到", "", "", extra.get("info", ""))
                not_found += 1

            done += 1
            self.progress.emit(done, total)
            if done % 50 == 0:
                self.log.emit(f"进度 {done}/{total}  找到:{found} 未找到:{not_found} 错误:{error}")
            time.sleep(0.5)

        self.finished_signal.emit(found, not_found, error, skip)


class WorkbenchWorker(QThread):
    log        = pyqtSignal(str)
    batch_done = pyqtSignal(int, int, int)   # batch_idx, total_batches, count
    all_done   = pyqtSignal()

    def __init__(self, objects, batch_size=200):
        super().__init__()
        self.objects    = objects
        self.batch_size = batch_size
        self._stop      = False
        self._go        = False   # 等待继续信号

    def stop(self):
        self._stop = True
        self._go   = True

    def continue_next(self):
        self._go = True

    def run(self):
        total_b = math.ceil(len(self.objects) / self.batch_size)
        self.log.emit(f"共 {len(self.objects)} 个，分 {total_b} 批（每批 {self.batch_size}）")

        for bi in range(total_b):
            if self._stop: break
            batch = self.objects[bi * self.batch_size:(bi + 1) * self.batch_size]
            self.log.emit(f"\n--- 第 {bi+1}/{total_b} 批（{len(batch)} 个）---")
            try:
                cnt = api_core.set_workbench_via_cdp(batch)
                self.log.emit(f"写入工作台: {cnt} 个")
                self.batch_done.emit(bi + 1, total_b, len(batch))
            except Exception as e:
                self.log.emit(f"写入失败: {e}")
                continue

            if bi < total_b - 1:
                self._go = False
                while not self._go and not self._stop:
                    time.sleep(0.2)

        if not self._stop:
            self.all_done.emit()


# ─────────────────────────── 主窗口 ──────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("视频素材批量查询工具")
        self.setMinimumSize(1200, 800)

        self.df             = None
        self.df_filtered    = None
        self.excel_path     = ""
        self.token          = ""
        self.sheet_names    = []
        self.current_sheet  = None
        self.search_worker  = None
        self.wb_worker      = None
        self._wb_pushed     = 0
        self._wb_total_objs = 0
        self.results_dir    = RESULTS_DIR

        self._build_ui()
        self._connect_token()

    # ──────────── UI 构建 ────────────

    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)
        vbox = QVBoxLayout(root)
        vbox.setContentsMargins(10, 10, 10, 6)
        vbox.setSpacing(8)

        def _label(text, tip=None, bold=False):
            lb = QLabel(text)
            if bold:
                lb.setStyleSheet("font-weight:bold")
            if tip:
                lb.setToolTip(tip)
            return lb

        def _btn(text, color, hover, slot, tip=None, width=None):
            b = QPushButton(text)
            b.setFixedHeight(36)
            if width:
                b.setFixedWidth(width)
            b.setStyleSheet(
                f"QPushButton{{background:{color};color:white;font-size:13px;"
                f"border-radius:5px;padding:0 14px}}"
                f"QPushButton:hover{{background:{hover}}}"
                f"QPushButton:disabled{{background:#bdbdbd;color:#eee}}"
            )
            b.clicked.connect(slot)
            if tip:
                b.setToolTip(tip)
            return b

        # ═══════════════════════════════════════════
        # 第一区：文件导入 & 设置
        # ═══════════════════════════════════════════
        fg = QGroupBox("第一步：导入表格文件")
        fg.setStyleSheet("QGroupBox{font-weight:bold;font-size:13px;padding-top:6px}"
                         "QGroupBox::title{subcontrol-origin:margin;left:10px}")
        fg_layout = QVBoxLayout(fg)
        fg_layout.setSpacing(6)

        # 第一行：文件选择 + 工作表 + 搜索列
        fl = QHBoxLayout()
        fl.setSpacing(8)

        btn_open = QPushButton("  选择文件")
        btn_open.setFixedHeight(34)
        btn_open.setFixedWidth(100)
        btn_open.setStyleSheet(
            "QPushButton{background:#1976D2;color:white;font-size:13px;"
            "border-radius:5px;font-weight:bold}"
            "QPushButton:hover{background:#1565C0}"
        )
        btn_open.setToolTip("支持 Excel (.xlsx/.xls)、CSV、Apple Numbers (.numbers)")
        btn_open.clicked.connect(self._open_file)

        self.file_label = QLabel("未选择文件  （支持 .xlsx / .xls / .csv / .numbers）")
        self.file_label.setStyleSheet("color:#555; font-size:12px")
        self.file_label.setMinimumWidth(280)

        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(110)
        self.sheet_combo.setToolTip("选择要查询的工作表（Sheet）")
        self.sheet_combo.currentIndexChanged.connect(self._on_sheet_changed)

        self.col_combo = QComboBox()
        self.col_combo.setMinimumWidth(200)
        self.col_combo.setToolTip("选择「素材名称」所在列，程序将逐行取该列值查询接口")

        self.max_count_combo = QComboBox()
        self.max_count_combo.setFixedWidth(130)
        self.max_count_combo.addItems(["每条最多 60 个", "每条最多 120 个", "每条最多 300 个", "每条获取全部"])
        self.max_count_combo.setToolTip(
            "接口为模糊搜索，一个关键词可能匹配多条视频：\n"
            "• 最多 60 个：只取第一页，速度最快\n"
            "• 最多 120/300 个：自动翻页，结果更全\n"
            "• 获取全部：翻完所有页，关键词范围大时较慢"
        )

        fl.addWidget(btn_open)
        fl.addWidget(self.file_label, 1)
        fl.addWidget(_label("工作表:", "选择 Sheet")); fl.addWidget(self.sheet_combo)
        fl.addWidget(_label("素材名称列:", "选择包含素材名称的列")); fl.addWidget(self.col_combo)
        fl.addWidget(_label("匹配数量:")); fl.addWidget(self.max_count_combo)
        fg_layout.addLayout(fl)

        # 第二行：表头设置 + 起始行
        fl2 = QHBoxLayout()
        fl2.setSpacing(16)

        self.chk_header = QCheckBox("首行是表头（列名行）")
        self.chk_header.setChecked(True)
        self.chk_header.setToolTip(
            "勾选（推荐）：表格第一行是列名，如「素材名称、素材ID...」\n"
            "不勾选：表格第一行就是数据，没有列名行\n\n"
            "切换后会自动重新加载表格"
        )
        self.chk_header.stateChanged.connect(self._on_header_changed)

        self.start_spin = QSpinBox()
        self.start_spin.setPrefix("从第 ")
        self.start_spin.setSuffix(" 行开始查询")
        self.start_spin.setMinimum(0)
        self.start_spin.setMaximum(99999)
        self.start_spin.setFixedWidth(160)
        self.start_spin.setToolTip(
            "断点续查 / 跳过指定行数：\n"
            "• 0（默认）= 从头开始查全部数据行\n"
            "• 有表头时：行号从 1 开始（第1行=第一条数据）\n"
            "• 无表头时：行号从 1 开始（第1行=第一行数据）\n"
            "• 例如填 50 = 跳过前 50 条，从第 51 条开始"
        )

        self.header_hint = QLabel("")
        self.header_hint.setStyleSheet("color:#888; font-size:11px")
        self._update_header_hint()

        fl2.addWidget(self.chk_header)
        fl2.addWidget(self.start_spin)
        fl2.addWidget(self.header_hint)
        fl2.addStretch()
        fg_layout.addLayout(fl2)

        vbox.addWidget(fg)

        # ═══════════════════════════════════════════
        # 第二区：操作按钮
        # ═══════════════════════════════════════════
        op_group = QGroupBox("第二步：查询操作")
        op_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:13px;padding-top:6px}"
                               "QGroupBox::title{subcontrol-origin:margin;left:10px}")
        op_layout = QHBoxLayout(op_group)
        op_layout.setSpacing(8)

        self.btn_search = _btn(
            "▶  开始查询", "#43A047", "#2E7D32", self._start_search,
            tip="按表格逐行查询素材，查询中可随时点「暂停」")
        self.btn_stop = _btn(
            "⏹  暂停", "#e53935", "#b71c1c", self._stop_search,
            tip="停止当前查询，已查结果不丢失，下次可从断点继续")
        self.btn_stop.setEnabled(False)

        self.btn_requery_errors = _btn(
            "重查错误行", "#8E24AA", "#6A1B9A", self._requery_errors,
            tip="将状态为「错误」的行重置并重新查询")
        self.btn_reset_all = _btn(
            "重置全部重查", "#6D4C41", "#4E342E", self._reset_all,
            tip="清空所有行的查询状态，从头开始重新查询全部行")

        sep1 = QLabel("  |  ")
        sep1.setStyleSheet("color:#ccc; font-size:18px")

        self.btn_workbench = _btn(
            "加入工作台", "#1E88E5", "#1565C0", self._add_to_workbench,
            tip="将所有「找到」状态的视频 ID 批量推送到素材网工作台\n每批最多 200 个，超出自动分批并等待确认")

        sep2 = QLabel("  |  ")
        sep2.setStyleSheet("color:#ccc; font-size:18px")

        self.btn_save = _btn(
            "保存结果", "#FB8C00", "#E65100", self._save_excel,
            tip="将当前表格（含查询结果列）另存为 Excel 文件")
        self.btn_token = _btn(
            "刷新登录", "#546E7A", "#37474F", self._refresh_token,
            tip="Token 过期或查询报「请登录」时点此\n程序会自动打开浏览器，在浏览器登录后自动获取 Token")

        for w in [self.btn_search, self.btn_stop,
                  self.btn_requery_errors, self.btn_reset_all,
                  sep1,
                  self.btn_workbench,
                  sep2,
                  self.btn_save, self.btn_token]:
            op_layout.addWidget(w)
        op_layout.addStretch()

        # Token 状态指示灯
        self.token_indicator = QLabel("● Token 未就绪")
        self.token_indicator.setStyleSheet("color:#e53935; font-size:12px; font-weight:bold")
        op_layout.addWidget(self.token_indicator)

        vbox.addWidget(op_group)

        # ═══════════════════════════════════════════
        # 第三区：筛选 & 结果目录
        # ═══════════════════════════════════════════
        filter_group = QGroupBox("第三步：查看结果")
        filter_group.setStyleSheet("QGroupBox{font-weight:bold;font-size:13px;padding-top:6px}"
                                   "QGroupBox::title{subcontrol-origin:margin;left:10px}")
        filter_layout = QHBoxLayout(filter_group)
        filter_layout.setSpacing(8)

        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["全部行", "找到", "未找到", "错误", "未查询"])
        self.filter_combo.setFixedWidth(90)
        self.filter_combo.setToolTip(
            "按查询状态筛选表格行：\n"
            "• 找到：接口有返回结果\n"
            "• 未找到：接口返回空\n"
            "• 错误：网络或 Token 问题\n"
            "• 未查询：尚未查询的行"
        )
        self.filter_combo.currentIndexChanged.connect(self._apply_filter)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("在表格中搜索素材名称关键词...")
        self.search_box.setFixedWidth(240)
        self.search_box.setToolTip("输入关键词实时过滤表格，不影响查询数据")
        self.search_box.textChanged.connect(self._apply_filter)

        self.row_count_label = QLabel("显示: 0 / 0 行")
        self.row_count_label.setStyleSheet("color:#555; font-size:12px")

        filter_layout.addWidget(_label("状态筛选:"))
        filter_layout.addWidget(self.filter_combo)
        filter_layout.addWidget(_label("  名称搜索:"))
        filter_layout.addWidget(self.search_box)
        filter_layout.addWidget(self.row_count_label)
        filter_layout.addStretch()

        filter_layout.addWidget(_label("结果目录:"))
        self.dir_label = QLabel(RESULTS_DIR)
        self.dir_label.setStyleSheet("color:#1565C0; font-size:11px")
        self.dir_label.setMaximumWidth(300)
        self.dir_label.setToolTip(RESULTS_DIR)
        btn_dir = QPushButton("更改")
        btn_dir.setFixedWidth(50)
        btn_dir.setFixedHeight(28)
        btn_dir.setToolTip("更改结果 Excel 的保存目录")
        btn_dir.clicked.connect(self._choose_results_dir)
        filter_layout.addWidget(self.dir_label)
        filter_layout.addWidget(btn_dir)

        vbox.addWidget(filter_group)

        # ═══════════════════════════════════════════
        # 表格 + 日志（可拖拽分隔）
        # ═══════════════════════════════════════════
        splitter = QSplitter(Qt.Orientation.Vertical)

        self.table = QTableWidget()
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectItems)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.table.setSortingEnabled(True)
        self.table.setStyleSheet(
            "QTableWidget{gridline-color:#e0e0e0; font-size:12px}"
            "QHeaderView::section{background:#f5f5f5;font-weight:bold;padding:4px;"
            "border:1px solid #ddd}"
        )
        self.table.keyPressEvent = self._table_key_press
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._table_context_menu)
        splitter.addWidget(self.table)

        log_g = QGroupBox("运行日志  （实时显示查询进度和错误信息）")
        log_g.setStyleSheet("QGroupBox{font-size:12px;color:#555}")
        log_l = QVBoxLayout(log_g)
        log_l.setContentsMargins(4, 4, 4, 4)
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        self.log_text.setFont(QFont("Menlo", 11))
        self.log_text.setMaximumHeight(160)
        self.log_text.setStyleSheet("background:#1e1e1e; color:#d4d4d4; border:none")
        log_l.addWidget(self.log_text)
        splitter.addWidget(log_g)

        splitter.setStretchFactor(0, 5)
        splitter.setStretchFactor(1, 1)
        vbox.addWidget(splitter, 1)

        # ═══════════════════════════════════════════
        # 底部：统计 + 进度条
        # ═══════════════════════════════════════════
        bottom = QHBoxLayout()
        self.stats_label = QLabel("找到: 0 行 / 0 个视频  |  未找到: 0  |  错误: 0  |  未查: 0")
        self.stats_label.setStyleSheet("font-size:12px; font-weight:bold; color:#333")

        self.wb_stats = QLabel("工作台：尚未推送")
        self.wb_stats.setStyleSheet("color:#1565C0; font-size:12px; font-weight:bold")

        bottom.addWidget(self.stats_label)
        bottom.addStretch()
        bottom.addWidget(self.wb_stats)
        vbox.addLayout(bottom)

        self.progress = QProgressBar()
        self.progress.setFixedHeight(20)
        self.progress.setStyleSheet(
            "QProgressBar{border:1px solid #ccc;border-radius:4px;background:#f5f5f5;text-align:center}"
            "QProgressBar::chunk{background:#43A047;border-radius:4px}"
        )
        vbox.addWidget(self.progress)

        self.status_bar = QStatusBar()
        self.status_bar.setStyleSheet("font-size:12px")
        self.setStatusBar(self.status_bar)

    # ──────────── Token ────────────

    def _choose_results_dir(self):
        d = QFileDialog.getExistingDirectory(
            self, "选择结果保存目录", self.results_dir)
        if d:
            self.results_dir = d
            self.dir_label.setText(d)
            self._log(f"结果目录已更改: {d}")

    def _connect_token(self):
        self._log("正在获取 Token...")
        self.status_bar.showMessage("正在获取 Token...")
        self._token_worker = TokenWorker(ask_login=False)
        self._token_worker.log.connect(self._log)
        self._token_worker.success.connect(self._on_token_ok)
        self._token_worker.failed.connect(self._on_token_fail_silent)
        self._token_worker.start()

    def _refresh_token(self):
        self._log("正在关闭浏览器并以调试模式重新启动...")
        self.btn_token.setEnabled(False)
        self._token_worker = TokenWorker(ask_login=True)
        self._token_worker.log.connect(self._log)
        self._token_worker.success.connect(self._on_token_ok)
        self._token_worker.failed.connect(self._on_token_fail)
        self._token_worker.start()

    def _on_token_ok(self, token):
        self.token = token
        self._log("Token 获取成功，可以开始查询")
        self.status_bar.showMessage("Token 就绪 ✓")
        self.btn_token.setEnabled(True)
        self.token_indicator.setText("● Token 已就绪")
        self.token_indicator.setStyleSheet("color:#43A047; font-size:12px; font-weight:bold")

    def _on_token_fail_silent(self, err):
        self._log(f"自动获取 Token 失败，请点「刷新登录」按钮后在浏览器中登录: {err}")
        self.status_bar.showMessage("请点「刷新登录」在浏览器中登录")
        self.btn_token.setEnabled(True)
        self.token_indicator.setText("● Token 未就绪  点「刷新登录」")
        self.token_indicator.setStyleSheet("color:#e53935; font-size:12px; font-weight:bold")

    def _on_token_fail(self, err):
        self._log(f"Token 获取失败: {err}")
        self.status_bar.showMessage("Token 获取失败")
        self.btn_token.setEnabled(True)
        self.token_indicator.setText("● Token 获取失败  点「刷新登录」")
        self.token_indicator.setStyleSheet("color:#e53935; font-size:12px; font-weight:bold")
        QMessageBox.warning(self, "登录失败", f"获取失败，请确认浏览器已打开并登录\n\n{err}")

    # ──────────── 日志 ────────────

    def _table_context_menu(self, pos):
        menu = QMenu(self)
        act_copy_cell  = menu.addAction("复制单元格")
        act_copy_row   = menu.addAction("复制整行")
        act_copy_all   = menu.addAction("复制选中区域")
        action = menu.exec(self.table.viewport().mapToGlobal(pos))
        if action == act_copy_cell:
            item = self.table.itemAt(pos)
            if item:
                QApplication.clipboard().setText(item.text())
        elif action == act_copy_row:
            row = self.table.rowAt(pos.y())
            if row >= 0:
                parts = []
                for ci in range(self.table.columnCount()):
                    it = self.table.item(row, ci)
                    parts.append(it.text() if it else "")
                QApplication.clipboard().setText("\t".join(parts))
        elif action == act_copy_all:
            self._copy_selection()

    def _copy_selection(self):
        selected = self.table.selectedRanges()
        if not selected:
            return
        sel_rows, sel_cols = set(), set()
        for r in selected:
            for ri in range(r.topRow(), r.bottomRow() + 1):
                sel_rows.add(ri)
            for ci in range(r.leftColumn(), r.rightColumn() + 1):
                sel_cols.add(ci)
        sel_rows, sel_cols = sorted(sel_rows), sorted(sel_cols)
        if len(sel_rows) == 1 and len(sel_cols) == 1:
            item = self.table.item(sel_rows[0], sel_cols[0])
            QApplication.clipboard().setText(item.text() if item else "")
            return
        headers = [self.table.horizontalHeaderItem(c).text()
                   if self.table.horizontalHeaderItem(c) else str(c)
                   for c in sel_cols]
        lines = ["\t".join(headers)]
        for ri in sel_rows:
            lines.append("\t".join(
                self.table.item(ri, ci).text() if self.table.item(ri, ci) else ""
                for ci in sel_cols))
        QApplication.clipboard().setText("\n".join(lines))

    def _table_key_press(self, event):
        is_copy = (event.key() == Qt.Key.Key_C and (
            event.modifiers() & Qt.KeyboardModifier.ControlModifier or
            event.modifiers() & Qt.KeyboardModifier.MetaModifier
        ))
        if is_copy:
            self._copy_selection()
        else:
            QTableWidget.keyPressEvent(self.table, event)

    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{ts}] {msg}")
        sb = self.log_text.verticalScrollBar()
        sb.setValue(sb.maximum())
        QApplication.processEvents()

    # ──────────── 文件 / Sheet ────────────

    def _open_file(self):
        dialog = QFileDialog(self, "打开 Excel / CSV / Numbers")
        dialog.setNameFilters([
            "表格文件 (*.xlsx *.xls *.csv *.numbers)",
            "Excel 文件 (*.xlsx *.xls)",
            "CSV 文件 (*.csv)",
            "Numbers 文件 (*.numbers)",
            "所有文件 (*)",
        ])
        dialog.setOption(QFileDialog.Option.DontUseNativeDialog, True)
        if dialog.exec() != QFileDialog.DialogCode.Accepted:
            return
        files = dialog.selectedFiles()
        if not files:
            return
        path = files[0]
        if not path:
            return
        self.excel_path = path
        self.file_label.setText(os.path.basename(path))
        try:
            if path.lower().endswith(".csv"):
                # CSV 文件：单 sheet
                self.sheet_names = ["Sheet1"]
                self.sheet_combo.blockSignals(True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(self.sheet_names)
                self.sheet_combo.blockSignals(False)
                self._load_sheet("Sheet1")
            elif path.lower().endswith(".numbers"):
                # Apple Numbers 文件
                sheets_data = _read_numbers_file(path)
                self._numbers_data = sheets_data
                self.sheet_names = list(sheets_data.keys())
                self.sheet_combo.blockSignals(True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(self.sheet_names)
                self.sheet_combo.blockSignals(False)
                self._load_sheet(self.sheet_names[0])
            else:
                # 先尝试 openpyxl，失败则 xlrd
                try:
                    xls = pd.ExcelFile(path, engine="openpyxl")
                except Exception:
                    xls = pd.ExcelFile(path, engine="xlrd")
                self.sheet_names = xls.sheet_names
                self.sheet_combo.blockSignals(True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(self.sheet_names)
                self.sheet_combo.blockSignals(False)
                self._auto_detect_header(path, self.sheet_names[0])
                self._load_sheet(self.sheet_names[0])
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"无法读取文件:\n{path}\n\n{e}")

    def _auto_detect_header(self, path, sheet):
        """检测第一行是否包含表头关键词，自动设置复选框"""
        if path.lower().endswith(".csv"):
            return
        try:
            try:
                df_raw = pd.read_excel(path, sheet_name=sheet, engine="openpyxl", header=None, nrows=1)
            except Exception:
                df_raw = pd.read_excel(path, sheet_name=sheet, engine="xlrd", header=None, nrows=1)
            first_row_vals = [str(v).strip() for v in df_raw.iloc[0].tolist()]
            header_keywords = {"素材名称", "查询状态", "视频ID", "视频id", "名称", "状态"}
            has_header = bool(header_keywords & set(first_row_vals))
            self.chk_header.blockSignals(True)
            self.chk_header.setChecked(has_header)
            self.chk_header.blockSignals(False)
            self._log(f"自动检测: {'有' if has_header else '无'}表头")
        except Exception:
            pass

    def _on_sheet_changed(self, idx):
        if 0 <= idx < len(self.sheet_names):
            self._load_sheet(self.sheet_names[idx])

    def _on_header_changed(self):
        """首行是否表头 切换时重新加载"""
        self._update_header_hint()
        if self.excel_path and self.current_sheet:
            self._load_sheet(self.current_sheet)

    def _update_header_hint(self):
        if not hasattr(self, "header_hint"):
            return
        if self.chk_header.isChecked():
            self.header_hint.setText("第一行作为列名，数据从第二行开始")
        else:
            self.header_hint.setText("无列名，第一行就是数据，列名自动生成为「列1、列2...」")

    def _load_sheet(self, sheet):
        try:
            has_header = self.chk_header.isChecked()
            header_arg = 0 if has_header else None
            if self.excel_path.lower().endswith(".csv"):
                self.df = pd.read_csv(self.excel_path, header=header_arg, encoding="utf-8-sig")
            elif self.excel_path.lower().endswith(".numbers"):
                raw_df = getattr(self, "_numbers_data", {}).get(sheet, pd.DataFrame())
                self.df = raw_df.reset_index(drop=True).copy()
                # numbers 文件已有表头，has_header 强制视为 True
                has_header = True
            else:
                try:
                    self.df = pd.read_excel(self.excel_path, sheet_name=sheet,
                                            engine="openpyxl", header=header_arg)
                except Exception:
                    self.df = pd.read_excel(self.excel_path, sheet_name=sheet,
                                            engine="xlrd", header=header_arg)

            if not has_header:
                # 无表头：生成「列1 / 列2 / ...」列名
                self.df.columns = [f"列{i+1}" for i in range(len(self.df.columns))]
                # 清理旧结果列（如已追加过查询结果列）
                drop_cols = [c for c in self.df.columns
                             if not self.df.empty and
                             str(self.df[c].iloc[0]).strip() in
                             {"查询状态", "视频ID", "查询备注", "查询时间", "匹配名称"}]
                if drop_cols:
                    self.df.drop(columns=drop_cols, inplace=True)
                    self._log(f"检测到旧结果列 {drop_cols}，已移除")

            self.current_sheet = sheet
            for col in ["查询状态", "视频ID", "匹配名称", "查询备注", "查询时间"]:
                if col not in self.df.columns:
                    self.df[col] = ""
                else:
                    self.df[col] = self.df[col].fillna("").astype(str)

            self._log(f"加载 [{sheet}]: {len(self.df)} 行，{'有' if has_header else '无'}表头")
            self.start_spin.setMaximum(max(0, len(self.df) - 1))

            # 构建搜索列下拉：有表头显示列名，无表头显示「列N: 首行内容预览」
            self.col_combo.clear()
            result_cols = {"查询状态", "视频ID", "匹配名称", "查询备注", "查询时间"}
            for c in self.df.columns:
                if str(c) in result_cols:
                    continue  # 结果列不放进搜索列选项
                if has_header:
                    self.col_combo.addItem(str(c), userData=c)
                else:
                    # 取第一行内容作为预览
                    preview = str(self.df[c].iloc[0])[:20] if not self.df.empty else ""
                    self.col_combo.addItem(f"{c}: {preview}", userData=c)

            # 自动选中「素材名称」列，或内容最像素材名称的列
            best = self._guess_name_col(has_header)
            if best is not None:
                for i in range(self.col_combo.count()):
                    if self.col_combo.itemData(i) == best:
                        self.col_combo.setCurrentIndex(i)
                        break

            self._update_stats_label()
            self._apply_filter()
            self.status_bar.showMessage(f"[{sheet}] {len(self.df)} 行  ({'有表头' if has_header else '无表头，请在「搜索列」选择素材名称列'})")
        except Exception as e:
            QMessageBox.critical(self, "加载失败", str(e))

    def _guess_name_col(self, has_header):
        """猜测哪一列是素材名称列"""
        if self.df is None or self.df.empty:
            return None
        result_cols = {"查询状态", "视频ID", "匹配名称", "查询备注", "查询时间"}
        data_cols = [c for c in self.df.columns if str(c) not in result_cols]
        if not data_cols:
            return None
        # 有表头：优先匹配列名
        if has_header:
            keywords = ["素材名称", "素材", "名称", "name", "视频名", "标题"]
            for kw in keywords:
                for c in data_cols:
                    if kw in str(c).lower():
                        return c
        # 无表头 or 未命中：找内容最长、最像名称的列（字符串且平均长度>5）
        best_col, best_score = None, 0
        for c in data_cols:
            sample = self.df[c].dropna().astype(str).head(10)
            avg_len = sample.str.len().mean() if len(sample) > 0 else 0
            # 像名称的特征：含中文、含连字符、长度适中
            has_chinese = sample.str.contains(r'[\u4e00-\u9fff]', regex=True).mean()
            score = avg_len * 0.5 + has_chinese * 10
            if score > best_score:
                best_score = score
                best_col = c
        return best_col

    # ──────────── 筛选 ────────────

    def _apply_filter(self):
        if self.df is None:
            return
        status_filter = self.filter_combo.currentText()
        keyword = self.search_box.text().strip().lower()
        name_col = self.col_combo.currentData() or self.col_combo.currentText()

        mask = pd.Series([True] * len(self.df), index=self.df.index)

        if status_filter == "未查询":
            mask &= self.df["查询状态"].astype(str).str.strip() == ""
        elif status_filter not in ("全部行", "全部"):
            mask &= self.df["查询状态"].astype(str).str.strip() == status_filter

        if keyword and name_col in self.df.columns:
            mask &= self.df[name_col].astype(str).str.lower().str.contains(keyword, na=False)

        self.df_filtered = self.df[mask].copy()
        self._render_table(self.df_filtered)
        self.row_count_label.setText(f"显示: {len(self.df_filtered)} / {len(self.df)} 行")

    def _render_table(self, df):
        self.table.setSortingEnabled(False)
        self.table.setRowCount(len(df))
        self.table.setColumnCount(len(df.columns))
        self.table.setHorizontalHeaderLabels([str(c) for c in df.columns])

        name_col = self.col_combo.currentData() or self.col_combo.currentText()
        key_cols = {name_col, "查询状态", "视频ID", "匹配名称", "查询备注", "查询时间"}

        color_map = {"找到": "#C8E6C9", "未找到": "#FFF9C4", "错误": "#FFCDD2"}

        for ri, (_, row) in enumerate(df.iterrows()):
            status = str(row.get("查询状态", "")).strip()
            row_color = color_map.get(status)
            for ci, col in enumerate(df.columns):
                val = row[col]
                text = "" if pd.isna(val) else str(val)
                if str(col) not in key_cols and len(text) > 25:
                    text = text[:25] + "…"
                item = QTableWidgetItem(text)
                item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                if str(col) == "查询状态" and row_color:
                    item.setBackground(QColor(row_color))
                elif row_color:
                    item.setBackground(QColor(row_color).lighter(120))
                self.table.setItem(ri, ci, item)

        self.table.resizeColumnsToContents()
        self.table.setSortingEnabled(True)

    def _update_row(self, orig_idx, name, status, ids, matched_names, remark):
        if self.df is None: return
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.df.at[orig_idx, "查询状态"]  = status
        self.df.at[orig_idx, "视频ID"]    = ids
        self.df.at[orig_idx, "匹配名称"]  = matched_names
        self.df.at[orig_idx, "查询备注"]  = remark
        self.df.at[orig_idx, "查询时间"]  = now
        # 实时更新统计标签
        self._update_stats_label()
        # 在表格中找到对应行并原地更新（不重渲染整表）
        self._patch_table_row(orig_idx, status, ids, matched_names, remark, now)

    def _patch_table_row(self, orig_idx, status, ids, matched_names, remark, now):
        """在当前表格中找到 orig_idx 对应的行，原地更新状态/ID/备注"""
        if self.df_filtered is None:
            return
        color_map = {"找到": "#C8E6C9", "未找到": "#FFF9C4", "错误": "#FFCDD2"}
        row_color = color_map.get(status)
        cols = list(self.df_filtered.columns) if self.df_filtered is not None else list(self.df.columns)
        update_vals = {
            "查询状态": status, "视频ID": ids,
            "匹配名称": matched_names, "查询备注": remark, "查询时间": now,
        }
        # 找到表格中 orig_idx 对应的显示行（df_filtered 的 index）
        if self.df_filtered is not None and orig_idx in self.df_filtered.index:
            display_row = list(self.df_filtered.index).index(orig_idx)
            for ci, col in enumerate(cols):
                col_s = str(col)
                if col_s in update_vals:
                    val = update_vals[col_s]
                    item = self.table.item(display_row, ci)
                    if item is None:
                        item = QTableWidgetItem()
                        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
                        self.table.setItem(display_row, ci, item)
                    item.setText(str(val) if val else "")
                    if col_s == "查询状态" and row_color:
                        item.setBackground(QColor(row_color))
                    elif row_color:
                        item.setBackground(QColor(row_color).lighter(120))
        else:
            pass  # 该行不在当前筛选视图中，不需要更新表格

    def _update_stats_label(self):
        if self.df is None: return
        vc = self.df["查询状态"].value_counts()
        found    = vc.get("找到", 0)
        notfound = vc.get("未找到", 0)
        error    = vc.get("错误", 0)
        pending  = (self.df["查询状态"].astype(str).str.strip() == "").sum()
        # 计算视频总数（视频ID列逗号分隔）
        total_videos = 0
        if "视频ID" in self.df.columns:
            for v in self.df["视频ID"].dropna():
                s = str(v).strip()
                if s and s != "nan":
                    total_videos += len([x for x in s.split(",") if x.strip()])
        self.stats_label.setText(
            f"✅ 找到:{found}行/{total_videos}个视频  ❌ 未找到:{notfound}  ⚠️ 错误:{error}  ⏳ 未查:{pending}"
        )
        self.stats_label.setStyleSheet(
            "font-weight:bold; font-size:13px;"
            f"color:{'#1B5E20' if found > 0 else '#333'}"
        )

    # ──────────── 查询 ────────────

    def _start_search(self):
        if self.df is None:
            QMessageBox.warning(self, "提示", "请先导入 Excel"); return
        name_col = self.col_combo.currentData() or self.col_combo.currentText()
        if not name_col or name_col not in self.df.columns:
            QMessageBox.warning(self, "提示", "请选择搜索列"); return
        if not self.token:
            self._refresh_token()
            if not self.token: return

        self.btn_search.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.progress.setMaximum(len(self.df))
        self.progress.setValue(0)

        max_count_map = {"每条最多 60 个": 60, "每条最多 120 个": 120, "每条最多 300 个": 300, "每条获取全部": -1}
        max_count = max_count_map.get(self.max_count_combo.currentText(), 60)
        self._log(f"开始查询 | 列:{name_col} | 起始行:{self.start_spin.value()} | 最大结果:{self.max_count_combo.currentText()}")

        self.search_worker = SearchWorker(
            self.df, name_col, self.start_spin.value(), self.token, max_count)
        self.search_worker.progress.connect(
            lambda c, t: self.progress.setValue(c))
        self.search_worker.row_done.connect(self._update_row)
        self.search_worker.log.connect(self._log)
        self.search_worker.token_refreshed.connect(
            lambda t: setattr(self, "token", t))
        self.search_worker.finished_signal.connect(self._search_done)
        self.search_worker.start()

    def _stop_search(self):
        if self.search_worker:
            self.search_worker.stop()
            self._log("正在停止...")

    def _search_done(self, found, not_found, error, skip):
        self.btn_search.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self._update_stats_label()
        self._log(f"查询完成 | 找到:{found} 未找到:{not_found} 错误:{error} 跳过:{skip}")
        self.status_bar.showMessage("查询完成")
        self._auto_save()

    def _requery_errors(self):
        if self.df is None: return
        mask = self.df["查询状态"].astype(str).str.strip() == "错误"
        count = mask.sum()
        if count == 0:
            QMessageBox.information(self, "提示", "没有错误行"); return
        self.df.loc[mask, "查询状态"] = ""
        self.df.loc[mask, "视频ID"]   = ""
        self.df.loc[mask, "查询备注"] = ""
        self._log(f"已重置 {count} 条错误行，开始重查...")
        self._apply_filter()
        self._start_search()

    def _reset_all(self):
        if self.df is None: return
        ret = QMessageBox.question(
            self, "确认",
            f"将清空所有 {len(self.df)} 行的查询状态，重新查询全部行。\n确认吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if ret != QMessageBox.StandardButton.Yes: return
        self.df["查询状态"] = ""
        self.df["视频ID"]   = ""
        self.df["匹配名称"] = ""
        self.df["查询备注"] = ""
        self.df["查询时间"] = ""
        self._log(f"已重置全部 {len(self.df)} 行，开始重查...")
        self._apply_filter()
        self._start_search()

    def _save_sheet_to_file(self, path, sheet_name, df):
        """保存当前 sheet 到文件，保留其他 sheet 不动"""
        from openpyxl import load_workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
        import openpyxl

        try:
            wb = load_workbook(path)
        except Exception:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        wb.save(path)

    def _auto_save(self):
        if not self.excel_path or not self.current_sheet:
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = os.path.splitext(os.path.basename(self.excel_path))[0]
        out_path = os.path.join(self.results_dir, f"{base}_结果_{ts}.xlsx")
        try:
            self.df.to_excel(out_path, index=False, engine="openpyxl")
            self._log(f"结果已保存: {out_path}")
        except Exception as e:
            self._log(f"自动保存失败: {e}")

    def _save_excel(self):
        if self.df is None: return
        base = os.path.splitext(os.path.basename(self.excel_path))[0] if self.excel_path else "结果"
        default_path = os.path.join(self.results_dir, f"{base}_结果.xlsx")
        path, _ = QFileDialog.getSaveFileName(self, "保存", default_path, "Excel (*.xlsx)")
        if not path: return
        try:
            self._save_sheet_to_file(path, self.current_sheet or "Sheet1", self.df)
            self._log(f"已保存: {path}")
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e))

    # ──────────── 工作台 ────────────

    def _add_to_workbench(self):
        if self.df is None:
            QMessageBox.warning(self, "提示", "请先导入并查询"); return

        all_ids = []
        for v in self.df[self.df["查询状态"].astype(str).str.strip() == "找到"]["视频ID"].dropna():
            for sid in str(v).split(","):
                sid = sid.strip()
                if sid and sid not in ("nan", ""):
                    try:
                        all_ids.append(str(int(float(sid))))
                    except (ValueError, OverflowError):
                        pass
        all_ids = list(dict.fromkeys(all_ids))

        if not all_ids:
            QMessageBox.information(self, "提示", "没有可添加的视频 ID"); return

        self._log(f"收集到 {len(all_ids)} 个唯一视频 ID，获取完整数据...")
        self.btn_workbench.setEnabled(False)
        QApplication.processEvents()

        try:
            objects = api_core.fetch_video_objects_by_ids(self.token, all_ids)
        except Exception as e:
            self._log(f"获取失败: {e}")
            self.btn_workbench.setEnabled(True)
            return

        if not objects:
            QMessageBox.information(self, "提示", "未获取到视频数据")
            self.btn_workbench.setEnabled(True)
            return

        self._wb_total_objs = len(objects)
        self._wb_pushed     = 0
        self._update_wb_stats(0, math.ceil(len(objects) / 200))
        self._log(f"获取到 {len(objects)} 个视频对象，开始分批写入...")

        self.wb_worker = WorkbenchWorker(objects, batch_size=200)
        self.wb_worker.log.connect(self._log)
        self.wb_worker.batch_done.connect(self._wb_batch_done)
        self.wb_worker.all_done.connect(self._wb_all_done)
        self.progress.setMaximum(math.ceil(len(objects) / 200))
        self.progress.setValue(0)
        self.wb_worker.start()

    def _wb_batch_done(self, bi, total_b, count):
        self._wb_pushed = bi
        remain_batches  = total_b - bi
        remain_videos   = remain_batches * 200  # 估算
        self.progress.setValue(bi)
        self._update_wb_stats(bi, total_b)
        self._log(f"第 {bi}/{total_b} 批（{count}个）写入完成")

        if bi < total_b:
            ret = QMessageBox.question(
                self, "工作台",
                f"✅ 第 {bi} 批（{count} 个）已加入工作台！\n\n"
                f"请在浏览器处理推送后，点 Yes 继续第 {bi+1} 批\n"
                f"剩余 {total_b - bi} 批",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if ret == QMessageBox.StandardButton.Yes:
                if self.wb_worker: self.wb_worker.continue_next()
            else:
                if self.wb_worker: self.wb_worker.stop()
                self.btn_workbench.setEnabled(True)
                self._log("已停止")

    def _wb_all_done(self):
        self.btn_workbench.setEnabled(True)
        total_b = math.ceil(self._wb_total_objs / 200)
        self._update_wb_stats(total_b, total_b)
        self.progress.setValue(self.progress.maximum())
        self._log("全部批次加入工作台完成！")
        QMessageBox.information(self, "完成", "全部视频已加入工作台！")

    def _update_wb_stats(self, pushed_batches, total_batches):
        pushed_videos  = min(pushed_batches * 200, self._wb_total_objs)
        remain_videos  = self._wb_total_objs - pushed_videos
        self.wb_stats.setText(
            f"工作台: 总计 {self._wb_total_objs} 个 | "
            f"已推 {pushed_batches}/{total_batches} 批 ({pushed_videos}个) | "
            f"剩余 {remain_videos} 个"
        )


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
